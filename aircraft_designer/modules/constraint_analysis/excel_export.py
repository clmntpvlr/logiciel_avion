"""Excel export for constraint analysis results."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.drawing.image import Image


class ConstraintExcelExporter:
    """Handle Excel export using openpyxl."""

    def __init__(self, path: Path) -> None:
        self.path = Path(path)

    def export(
        self,
        inputs: Dict,
        sweep: Dict,
        results: Dict,
        plot_path: Path | None = None,
    ) -> Path:
        wb = Workbook()
        self._write_inputs(wb, inputs, sweep)
        self._write_curves(wb, results["curves"])
        self._write_envelope(wb, results["envelope"])
        self._write_summary(wb, results)
        if plot_path and plot_path.is_file():
            ws = wb["Summary"]
            img = Image(str(plot_path))
            ws.add_image(img, "E2")
        wb.save(self.path)
        return self.path

    def _write_inputs(self, wb: Workbook, inputs: Dict, sweep: Dict) -> None:
        ws = wb.active
        ws.title = "Inputs"
        row = 1
        for section, data in inputs.items():
            ws.cell(row=row, column=1, value=section)
            row += 1
            for key, val in data.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=val)
                row += 1
            row += 1
        ws.cell(row=row, column=1, value="Sweep")
        for idx, key in enumerate(["ws_min", "ws_max", "ws_step", "units"], start=1):
            ws.cell(row=row + idx, column=1, value=key)
            ws.cell(row=row + idx, column=2, value=sweep.get(key))

    def _write_curves(self, wb: Workbook, curves: Dict[str, List[List[float]]]) -> None:
        ws = wb.create_sheet("Curves")
        col = 1
        for name, data in curves.items():
            ws.cell(row=1, column=col, value=f"{name}_ws")
            ws.cell(row=1, column=col + 1, value=f"{name}_tw")
            for i, (ws_val, tw_val) in enumerate(data, start=2):
                ws.cell(row=i, column=col, value=ws_val)
                ws.cell(row=i, column=col + 1, value=tw_val)
            col += 3

    def _write_envelope(self, wb: Workbook, envelope: List[List[float]]) -> None:
        ws = wb.create_sheet("Envelope")
        ws.cell(row=1, column=1, value="ws")
        ws.cell(row=1, column=2, value="tw")
        for i, (ws_val, tw_val) in enumerate(envelope, start=2):
            ws.cell(row=i, column=1, value=ws_val)
            ws.cell(row=i, column=2, value=tw_val)

    def _write_summary(self, wb: Workbook, results: Dict) -> None:
        ws = wb.create_sheet("Summary")
        rec = results.get("recommendation", {})
        ws.cell(row=1, column=1, value="WS_recommendé")
        ws.cell(row=1, column=2, value=rec.get("ws"))
        ws.cell(row=2, column=1, value="TW_recommandé")
        ws.cell(row=2, column=2, value=rec.get("tw"))


__all__ = ["ConstraintExcelExporter"]
